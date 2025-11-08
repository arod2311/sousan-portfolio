import pandas as pd
import os
import datetime
from openpyxl.styles import numbers, PatternFill

def process_ar_balance_file():
    # ----------------------------------------------------------------
    # 1) Define your input CSV and output folder
    # ----------------------------------------------------------------
    file_path = r"C:\Users\arodriguez\Documents\Projects\sousan\Production\arBalanceScript\Files\raw\arBalJan2025.csv"
    processed_folder = r"C:\Users\arodriguez\Documents\Projects\sousan\Production\arBalanceScript\Files\processed"

    # Construct output filename
    today = datetime.datetime.today()
    output_filename = (
        f"ProcessedARBalancewithMoneyBuckets"
        f"{today.month:02d}{today.day:02d}{today.year}.xlsx"
    )
    output_path = os.path.join(processed_folder, output_filename)

    print("Starting AR Balance script...")
    print("Reading file:", file_path)

    # ----------------------------------------------------------------
    # 2) Read CSV
    # ----------------------------------------------------------------
    df = pd.read_csv(file_path, header=None, dtype=str)
    print("Initial shape:", df.shape)

    # ----------------------------------------------------------------
    # 3) Remove columns by index slices
    # ----------------------------------------------------------------
    df.drop(df.columns[0:59], axis=1, inplace=True)
    print("After removing columns 0-59:", df.shape)

    df.drop(df.columns[12:], axis=1, inplace=True)
    print("After removing columns from index 12 onward:", df.shape)

    df.drop(df.columns[3:5], axis=1, inplace=True)
    df.drop(df.columns[3], axis=1, inplace=True)
    print("After additional column drops:", df.shape)

    # ----------------------------------------------------------------
    # 4) Remove rows with empty first column
    # ----------------------------------------------------------------
    df[df.columns[0]] = df[df.columns[0]].fillna('').str.strip()
    df = df[df[df.columns[0]] != '']
    print("After dropping rows missing col A:", df.shape)

    # ----------------------------------------------------------------
    # Remove duplicates based on col A (AR Code)
    # ----------------------------------------------------------------
    df.drop_duplicates(subset=[df.columns[0]], keep='first', inplace=True)
    print("After dropping duplicates on col A:", df.shape)

    # ----------------------------------------------------------------
    # 5) Rename columns to final 9
    # ----------------------------------------------------------------
    df.columns = [
        "AR Code",
        "Name",
        "Telephone",
        "Current",
        "31-60 Days",
        "61-90 Days",
        "91-120 Days",
        "Over 120 Days",
        "Total Due"
    ]
    print("Columns renamed. Current columns:", list(df.columns))

    # ----------------------------------------------------------------
    # Insert "Multiple Sites" column (index=9)
    # ----------------------------------------------------------------
    six_digits = df["AR Code"].str[:6]
    counts = six_digits.value_counts()
    df.insert(
        9,
        "Multiple Sites",
        six_digits.apply(lambda x: "Yes" if counts[x] > 1 else "No")
    )
    print('Added "Multiple Sites" column based on first 6 digits of AR Code.')

    # ----------------------------------------------------------------
    # Convert relevant columns (Current..Over 120 Days, Total Due) to numeric
    # ----------------------------------------------------------------
    def convert_currency_like(col_series):
        col_series = (col_series.fillna("")
                                .str.replace(r"\s+", "", regex=True)
                                .str.replace(r"\(", "-", regex=True)
                                .str.replace(r"\)", "", regex=True)
                                .str.replace(r"\$", "", regex=True)
                                .str.replace(",", "", regex=True))
        return pd.to_numeric(col_series, errors="coerce").fillna(0)

    for col_name in df.columns[3:8]:  # Current..Over 120 Days
        df[col_name] = convert_currency_like(df[col_name])
    df["Total Due"] = convert_currency_like(df["Total Due"])

    # ----------------------------------------------------------------
    # Subsets (buckets)
    # ----------------------------------------------------------------
    credits_df = df[df["Total Due"] < 0]
    under_100 = df[(df["Total Due"] >= 0) & (df["Total Due"] < 100)]
    between_100_200 = df[(df["Total Due"] >= 100) & (df["Total Due"] <= 200)]
    between_201_500 = df[(df["Total Due"] >= 201) & (df["Total Due"] <= 500)]
    between_501_1000 = df[(df["Total Due"] >= 501) & (df["Total Due"] <= 1000)]
    over_1000 = df[df["Total Due"] > 1000]

    print("Creating Excel workbook:", output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # READ ME
        read_me_data = {
            "Instructions": [
                "WELCOME TO THE AR BALANCE REPORT!",
                "",
                "WORKSHEET GUIDE:",
                "  * Summary: High-level overview of each bucket (counts, totals).",
                "  * Master File: All accounts (negatives, positives, zero).",
                "  * Credits Due: Negative 'Total Due' only.",
                "  * Under $100: Non-negative amounts below $100.",
                "  * $100-$200: Non-negative amounts between $100 and $200.",
                "  * $201-$500: Non-negative amounts between $201 and $500.",
                "  * $501-$1000: Non-negative amounts between $501 and $1000.",
                "  * > $1000: Non-negative amounts above $1000.",
                "",
                "NOTES:",
                "  * 'Multiple Sites' (last column) is 'Yes' if 'AR Code' first 6 digits appear on more than one row. Meaning Account has more than 1 site attached to Account #",
                "  * Parentheses in 'Total Due' indicate a negative (credit) balance.",
                "  * Columns D–H are numeric, so you can sum them in Excel.",
                "  * De-duplication occurs on 'AR Code'—only the first occurrence is kept.",
                "THANK YOU FOR USING THIS REPORT!"
            ]
        }
        read_me_df = pd.DataFrame(read_me_data)
        read_me_df.to_excel(writer, sheet_name="READ ME", index=False)

        # === NEW: SUMMARY WORKSHEET ===
        def summarize_sheet(name, data):
            return {
                "Worksheet": name,
                "Number of Accounts": data.shape[0],
                "Current": data["Current"].sum(),
                "31-60 Days": data["31-60 Days"].sum(),
                "61-90 Days": data["61-90 Days"].sum(),
                "91-120 Days": data["91-120 Days"].sum(),
                "Over 120 Days": data["Over 120 Days"].sum(),
                "Total Due": data["Total Due"].sum()
            }

        summary_rows = [
            summarize_sheet("Master File", df),
            summarize_sheet("Credits Due", credits_df),
            summarize_sheet("Under 100", under_100),
            summarize_sheet("100-200", between_100_200),
            summarize_sheet("201-500", between_201_500),
            summarize_sheet("501-1000", between_501_1000),
            summarize_sheet(">1000", over_1000),
        ]
        summary_df = pd.DataFrame(summary_rows)

        # Row 8: total sum excluding Master (rows 2..7)
        sum_excluding_master = summary_df.iloc[1:].sum(numeric_only=True)
        row_excluding_master = {
            "Worksheet": "Total Sum, Excluding Master",
            "Number of Accounts": int(sum_excluding_master["Number of Accounts"]),
            "Current": sum_excluding_master["Current"],
            "31-60 Days": sum_excluding_master["31-60 Days"],
            "61-90 Days": sum_excluding_master["61-90 Days"],
            "91-120 Days": sum_excluding_master["91-120 Days"],
            "Over 120 Days": sum_excluding_master["Over 120 Days"],
            "Total Due": sum_excluding_master["Total Due"],
        }
        summary_df = pd.concat([summary_df, pd.DataFrame([row_excluding_master])],
                               ignore_index=True)

        # Row 9: matching check, **per column**
        master_totals = summary_df.iloc[0]       # row 1
        excluding_master = summary_df.iloc[-1]   # row 8

        # We'll store YES/NO for each numeric column
        row_match = {"Worksheet": "Matching with Master?"}
        columns_to_check = [
            "Number of Accounts", "Current", "31-60 Days",
            "61-90 Days", "91-120 Days", "Over 120 Days", "Total Due"
        ]
        for col in columns_to_check:
            difference = master_totals[col] - excluding_master[col]
            row_match[col] = "YES" if abs(difference) < 0.00001 else "NO"

        summary_df = pd.concat([summary_df, pd.DataFrame([row_match])],
                               ignore_index=True)

        # Write summary to "Summary"
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Master File
        df.to_excel(writer, sheet_name="Master File", index=False)
        # Credits Due
        credits_df.to_excel(writer, sheet_name="Credits Due", index=False)
        # Under $100
        under_100.to_excel(writer, sheet_name="Under $100", index=False)
        # $100-$200
        between_100_200.to_excel(writer, sheet_name="$100-$200", index=False)
        # $201-$500
        between_201_500.to_excel(writer, sheet_name="$201-$500", index=False)
        # $501-$1000
        between_501_1000.to_excel(writer, sheet_name="$501-$1000", index=False)
        # > $1000
        over_1000.to_excel(writer, sheet_name="> $1000", index=False)

        # === 1) Apply Currency Formatting to columns D..H & "Total Due" in each sheet ===
        # We will do that for columns 4..8 + 9 in 1-based indexing => range(4,10).
        # That covers Current, 31-60, 61-90, 91-120, Over120, and TotalDue
        money_fmt = '"$"#,##0.00_);[Red]("$"#,##0.00)'

        def apply_currency_format_all_cols(sheet_name, num_rows):
            sheet = writer.sheets[sheet_name]
            for row in range(2, num_rows + 2):  # data rows
                for col in range(4, 10):        # columns D..I in Excel
                    cell = sheet.cell(row=row, column=col)
                    cell.number_format = money_fmt

        # Master File
        apply_currency_format_all_cols("Master File", df.shape[0])
        apply_currency_format_all_cols("Credits Due", credits_df.shape[0])
        apply_currency_format_all_cols("Under $100", under_100.shape[0])
        apply_currency_format_all_cols("$100-$200", between_100_200.shape[0])
        apply_currency_format_all_cols("$201-$500", between_201_500.shape[0])
        apply_currency_format_all_cols("$501-$1000", between_501_1000.shape[0])
        apply_currency_format_all_cols("> $1000", over_1000.shape[0])

        # === 2) Format columns C..H on the Summary sheet as currency (except col B is just count) ===
        # In the summary, A=Worksheet, B=# Accounts, C=Current, D=31-60, E=61-90, F=91-120, G=Over120, H=TotalDue
        summary_sheet = writer.sheets["Summary"]
        summary_num_rows = summary_df.shape[0]
        for row in range(2, summary_num_rows + 2):  # data rows, skip header
            for col in range(3, 9):  # columns C..H => 3..8
                cell = summary_sheet.cell(row=row, column=col)
                cell.number_format = money_fmt

        # === 3) Highlight "NO" cells in red on the last row of Summary (row = summary_num_rows + 1) ===
        # The last row has the "Matching with Master?" info. We'll highlight each "NO" cell in columns B..H
        #   (B=NumberOfAccounts, C..H are numeric columns)
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        last_row_index = summary_num_rows + 1  # 1-based indexing for Excel
        for col in range(2, 9):  # columns B..H in summary
            cell = summary_sheet.cell(row=last_row_index, column=col)
            if cell.value == "NO":
                cell.fill = red_fill

    print(f"Processing complete! Final Excel saved to: {output_path}")

if __name__ == "__main__":
    process_ar_balance_file()

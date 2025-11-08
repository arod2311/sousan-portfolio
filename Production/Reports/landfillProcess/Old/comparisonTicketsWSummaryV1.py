import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def main():
    # 1) Use today's date as part of filenames
    current_date = datetime.now().strftime("%Y-%m-%d")  # e.g. "2025-03-05"

    # 2) File paths (adjust as needed)
    reference_file = r"C:\Users\arodriguez\Documents\Projects\sousan\Production\landfillProcess\Files\rollOffCityOfLaredoTickets.csv"
    trux_file = rf"C:\Users\arodriguez\Documents\Projects\sousan\Production\landfillProcess\Files\Cleaned\Cleaned_Trux_DisposalTicketReport_{current_date}.csv"
    output_excel = rf"C:\Users\arodriguez\Documents\Projects\sousan\Production\landfillProcess\Files\Processed\RollOffComparisonSummary{current_date}.xlsx"

    # 3) If an old copy of the Excel file exists, delete it to ensure a fresh start
    if os.path.exists(output_excel):
        os.remove(output_excel)

    # 4) Create a new Workbook and rename the default sheet to "Summary"
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # ----- Load & Clean CSV Data -----
    reference_data = pd.read_csv(reference_file)
    trux_data = pd.read_csv(trux_file)

    # Strip column names
    reference_data.columns = reference_data.columns.str.strip()
    trux_data.columns = trux_data.columns.str.strip()

    # Identify columns
    reference_ticket_col = "Ticket"
    trux_ticket_col = "Ticket"
    reference_amount_col = "Amount"
    trux_amount_col = "Amount"
    found_in_trux_col = "Found in Trux"

    # Remove leading zeros + extra spaces from ticket columns
    reference_data[reference_ticket_col] = (
        reference_data[reference_ticket_col].astype(str).str.lstrip("0").str.strip()
    )
    trux_data[trux_ticket_col] = (
        trux_data[trux_ticket_col].astype(str).str.lstrip("0").str.strip()
    )

    # Convert the “Amount” columns to numeric
    reference_data[reference_amount_col] = (
        reference_data[reference_amount_col]
        .replace(r'[$,]', '', regex=True)
        .replace(r'^\s*-*\s*$', '0', regex=True)
        .astype(float)
        .fillna(0)
    )
    trux_data[trux_amount_col] = (
        trux_data[trux_amount_col]
        .replace(r'[$,]', '', regex=True)
        .replace(r'^\s*-*\s*$', '0', regex=True)
        .astype(float)
        .fillna(0)
    )

    # Format them back to currency strings
    reference_data[reference_amount_col] = reference_data[reference_amount_col].apply(lambda x: f"${x:,.2f}")
    trux_data[trux_amount_col] = trux_data[trux_amount_col].apply(lambda x: f"${x:,.2f}")

    # “Found in Trux” column
    reference_data[found_in_trux_col] = reference_data[reference_ticket_col].isin(
        trux_data[trux_ticket_col]
    ).map({True: "Yes", False: "No"})

    # ----- Generate Summary -----
    total_reference_tickets = len(reference_data)
    total_trux_tickets = len(trux_data)
    ticket_difference = total_reference_tickets - total_trux_tickets

    # Convert currency‐formatted strings back to floats for summation
    ref_total_charges = (
        reference_data[reference_amount_col]
        .replace(r'[$,]', '', regex=True)
        .astype(float)
        .sum()
    )
    trux_total_charges = (
        trux_data[trux_amount_col]
        .replace(r'[$,]', '', regex=True)
        .astype(float)
        .sum()
    )
    charge_difference = ref_total_charges - trux_total_charges

    tickets_found_in_trux = (reference_data[found_in_trux_col] == "Yes").sum()
    tickets_not_found_in_trux = (reference_data[found_in_trux_col] == "No").sum()

    summary_data = pd.DataFrame(
        [
            ["Total Charges", f"${ref_total_charges:,.2f}", f"${trux_total_charges:,.2f}", f"${charge_difference:,.2f}"],
            ["Total Tickets Quantities", total_reference_tickets, total_trux_tickets, ticket_difference],
            ["Tickets Found in Trux", tickets_found_in_trux, "", ""],
            ["Tickets Not Found in Trux", tickets_not_found_in_trux, "", ""],
        ],
        columns=["", "Landfill Invoice", "Trux Report (Currently)", "Difference"]
    )

    # ----- Write "Summary" data to the active sheet -----
    for row in dataframe_to_rows(summary_data, index=False, header=True):
        ws_summary.append(row)

    # Bold the header row
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)

    # Optionally format row 2 (the “Total Charges” row) with currency style
    for row in ws_summary.iter_rows(min_row=2, max_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = '"$"#,##0.00'

    # ----- Create and Write "Updated_rollOffCityOfLaredo" sheet -----
    ws_data = wb.create_sheet("UpdatedrollOffCityOfLaredo")
    for row in dataframe_to_rows(reference_data, index=False, header=True):
        ws_data.append(row)

    # Set the active sheet to "Summary" (optional)
    wb.active = 0

    # ----- Finally save the workbook -----
    wb.save(output_excel)

    # Console output
    print("\nSummary of Findings:")
    print(summary_data)
    print(f"\nData saved into: {output_excel}")

if __name__ == "__main__":
    main()

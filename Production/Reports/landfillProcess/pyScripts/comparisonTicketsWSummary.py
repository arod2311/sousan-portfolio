import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


def main():
    # 1) today's date for filenames
    current_date = datetime.now().strftime("%Y-%m-%d")

    # 2) base paths
    base_dir = r"C:\Users\arodriguez\Documents\repoProjects\sousan\Production\Reports\landfillProcess"
    reference_file = os.path.join(base_dir, "Files", "rollOffCityOfLaredoTickets.csv")
    trux_file = os.path.join(
        base_dir, "Files", "Cleaned",
        f"Cleaned_Trux_DisposalTicketReport_{current_date}.csv"
    )

    # 3) versioned Excel filename  (does NOT overwrite)
    base_name = f"RollOffComparisonSummary{current_date}"
    version   = 1
    while True:
        output_excel = os.path.join(
            base_dir, "Files", "Processed",
            f"{base_name}_V{version}.xlsx"
        )
        if not os.path.exists(output_excel):
            break
        version += 1                                    # V2, V3, …

    # CSV produced by addMatchColumn.py
    matched_csv = os.path.join(
        base_dir, "Files", "Processed","matchCsv",
        f"Updated_rollOffCityOfLaredo_WithMatch_{current_date}.csv"
    )

    # ------------------------------------------------------------------
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # ── Load & clean CSV data ─────────────────────────────────────────
    reference_data = pd.read_csv(reference_file)
    trux_data      = pd.read_csv(trux_file)

    reference_data.columns = reference_data.columns.str.strip()
    trux_data.columns      = trux_data.columns.str.strip()

    ticket_col = "Ticket"
    amt_col    = "Amount"

    # clean ticket & amount columns
    for df in (reference_data, trux_data):
        df[ticket_col] = df[ticket_col].astype(str).str.lstrip("0").str.strip()
        df[amt_col] = (
            df[amt_col]
            .replace(r'[$,]', '', regex=True)
            .replace(r'^\s*-*\s*$', '0', regex=True)
            .astype(float)
            .fillna(0)
            .apply(lambda x: f"${x:,.2f}")
        )

    # Found-in-Trux flag
    reference_data["Found in Trux"] = reference_data[ticket_col].isin(
        trux_data[ticket_col]
    ).map({True: "Yes", False: "No"})

    # ── Summary dataframe ─────────────────────────────────────────────
    totals_ref   = len(reference_data)
    totals_trux  = len(trux_data)
    charges_ref  = reference_data[amt_col].replace(r'[$,]', '', regex=True)\
                                          .astype(float).sum()
    charges_trux = trux_data[amt_col].replace(r'[$,]', '', regex=True)\
                                     .astype(float).sum()

    summary_df = pd.DataFrame(
        [
            ["Total Charges",
             f"${charges_ref:,.2f}",
             f"${charges_trux:,.2f}",
             f"${charges_ref - charges_trux:,.2f}"],
            ["Total Tickets Quantities",
             totals_ref, totals_trux, totals_ref - totals_trux],
            ["Tickets Found in Trux",
             (reference_data["Found in Trux"] == "Yes").sum(), "", ""],
            ["Tickets Not Found in Trux",
             (reference_data["Found in Trux"] == "No").sum(), "", ""],
        ],
        columns=["", "Landfill Invoice", "Trux Report (Currently)", "Difference"]
    )

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(row)

    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
    for row in ws_summary.iter_rows(min_row=2, max_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = '"$"#,##0.00'

    # ── Sheet 2: UpdatedrollOffCityOfLaredo  (original data) ──────────
    ws_ref = wb.create_sheet("UpdatedrollOffCityOfLaredo")
    for row in dataframe_to_rows(reference_data, index=False, header=True):
        ws_ref.append(row)

    # ── Sheet 3: TicketsWithMatch (if CSV exists) ─────────────────────
    if os.path.exists(matched_csv):
        match_df = pd.read_csv(matched_csv)
        ws_match = wb.create_sheet("TicketsWithMatch")
        for row in dataframe_to_rows(match_df, index=False, header=True):
            ws_match.append(row)

        # format Difference column red/black currency if present
        if "Difference" in match_df.columns:
            diff_idx = match_df.columns.get_loc("Difference") + 1  # 1-based idx
            cur_fmt = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            for r in ws_match.iter_rows(min_row=2, min_col=diff_idx, max_col=diff_idx):
                for cell in r:
                    txt = str(cell.value).replace("$", "").replace(",", "").strip()
                    if txt.startswith("(") and txt.endswith(")"):
                        txt = "-" + txt[1:-1]
                    try:
                        cell.value = float(txt)
                    except ValueError:
                        pass
                    cell.number_format = cur_fmt
    else:
        print(f"⚠️  Matched CSV not found: {matched_csv}")

    wb.active = 0
    wb.save(output_excel)
    print(f"\nWorkbook saved → {output_excel}")


if __name__ == "__main__":
    main()

import pandas as pd
import os
import re
import shutil

# For conditional formatting
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# ------------------------------------------------------------------------
# 1) Configuration
# ------------------------------------------------------------------------
ARCHIVE_FOLDER = (
    r"C:\Users\arodriguez\Documents\repoProjects\sousan\Production\Reports\yardageProductivity\files\Archive\2025\MonthlyReports"
)
MASTER_OUTPUT_FOLDER = (
    r"C:\Users\arodriguez\Documents\repoProjects\sousan\Production\Reports\yardageProductivity\files\MasterFile"
)
MASTER_FILENAME = "Master_Monthly_FL_Yardage_Report_Updated.xlsx"

os.makedirs(MASTER_OUTPUT_FOLDER, exist_ok=True)
MASTER_WORKBOOK_PATH = os.path.join(MASTER_OUTPUT_FOLDER, MASTER_FILENAME)

# Regex to detect months
month_pattern = re.compile(
    r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?"
    r"|Aug(?:ust)?|Sep(?:t)?(?:ember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\D?2025",
    re.IGNORECASE
)

# Map from month label -> (sheet_name, numeric_sort_index)
month_map = {
    "Jan": ("Jan-2025", 1),
    "January": ("Jan-2025", 1),
    "Feb": ("Feb-2025", 2),
    "February": ("Feb-2025", 2),
    "Mar": ("March-2025", 3),
    "March": ("March-2025", 3),
    "Apr": ("Apr-2025", 4),
    "April": ("Apr-2025", 4),
    "May": ("May-2025", 5),
    "Jun": ("Jun-2025", 6),
    "June": ("Jun-2025", 6),
    "Jul": ("Jul-2025", 7),
    "July": ("Jul-2025", 7),
    "Aug": ("Aug-2025", 8),
    "August": ("Aug-2025", 8),
    "Sep": ("Sep-2025", 9),
    "Sept": ("Sep-2025", 9),
    "September": ("Sep-2025", 9),
    "Oct": ("Oct-2025", 10),
    "October": ("Oct-2025", 10),
    "Nov": ("Nov-2025", 11),
    "November": ("Nov-2025", 11),
    "Dec": ("Dec-2025", 12),
    "December": ("Dec-2025", 12),
}

# Columns to compare month-over-month
NUMERIC_COLS = [
    "# of Services",
    "Total Yardage",
    "Size-0",
    "Size-2",
    "Size-4",
    "Size-6",
    "Size-8",
    "Size-10",
]

# ------------------------------------------------------------------------
# 2) The existing monthly file processor (unchanged)
#    (Identical to your current script, except consolidated for brevity)
# ------------------------------------------------------------------------
import pandas as pd

day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
day_index_map = {d: i for i, d in enumerate(day_order)}

def process_monthly_file(filepath):
    """
    Returns a DataFrame with columns:
        [RouteName, Day, # of Services, Total Yardage, Size-0, Size-2, Size-4, Size-6, Size-8, Size-10]
    sorted by (Route, Day).
    """
    df_all_data = pd.DataFrame()
    try:
        df_all = pd.read_excel(filepath, sheet_name="All Data")
    except ValueError:
        return df_all_data  # no All Data => empty

    if df_all.empty:
        return df_all_data

    # # of Services & Total Yardage
    services_df = (
        df_all.groupby(["RouteName", "Day"])
        .size()
        .reset_index(name="# of Services")
    )
    yardage_df = (
        df_all.groupby(["RouteName", "Day"])["Yardage"]
        .sum()
        .reset_index(name="Total Yardage")
    )
    merged_all = pd.merge(services_df, yardage_df, on=["RouteName", "Day"], how="outer")

    # Summary by Route-Day
    try:
        df_summary = pd.read_excel(filepath, sheet_name="Summary by Route-Day")
    except ValueError:
        df_summary = pd.DataFrame()

    if not df_summary.empty:
        pivot_sizes = df_summary.pivot_table(
            index=["RouteName", "Day"],
            columns="Size",
            values="Count",
            aggfunc="sum",
            fill_value=0
        ).reset_index()
    else:
        pivot_sizes = pd.DataFrame(columns=["RouteName", "Day"])

    desired_sizes = [0, 2, 4, 6, 8, 10]
    for c in pivot_sizes.columns:
        if c not in ("RouteName", "Day"):
            pivot_sizes.rename(columns={c: f"Size-{int(c)}"}, inplace=True)
    for sz in desired_sizes:
        szcol = f"Size-{sz}"
        if szcol not in pivot_sizes.columns:
            pivot_sizes[szcol] = 0

    merged = pd.merge(merged_all, pivot_sizes, on=["RouteName", "Day"], how="outer")

    # Fill NaN
    numeric_cols = ["# of Services", "Total Yardage"] + [f"Size-{s}" for s in desired_sizes]
    for nc in numeric_cols:
        merged[nc] = merged[nc].fillna(0)

    # Sorting
    import re
    def sort_key(route_name, day):
        route_type = 0 if "Route" in str(route_name) else 1
        m = re.search(r"(\d+)$", str(route_name))
        route_num = int(m.group(1)) if m else 999999
        day_num = day_index_map.get(day, 99)
        return (route_type, route_num, day_num)

    merged["sort_tuple"] = merged.apply(lambda r: sort_key(r["RouteName"], r["Day"]), axis=1)
    merged.sort_values(by="sort_tuple", inplace=True)
    merged.drop(columns=["sort_tuple"], inplace=True)

    final_cols = [
        "RouteName", "Day",
        "# of Services", "Total Yardage",
        "Size-0", "Size-2", "Size-4", "Size-6", "Size-8", "Size-10"
    ]
    df_all_data = merged[final_cols]
    return df_all_data

# ------------------------------------------------------------------------
# 3) Utility to compute difference & percent
# ------------------------------------------------------------------------
def compute_differences(current, previous):
    """
    Returns (diff, percent_float).
    If previous=0, we treat the percent as ±100% based on sign of diff.
    """
    diff = current - previous
    if abs(previous) < 1e-9:
        # Avoid division by zero
        if diff > 0:
            pct = 1.0   # +100% => decimal 1.0
        elif diff < 0:
            pct = -1.0  # -100% => decimal -1.0
        else:
            pct = 0.0
    else:
        # Remove "* 100"
        pct = diff / previous  # e.g. -0.264 means -26.4%

    return diff, pct

# ------------------------------------------------------------------------
# 4) Main builder: read + add "change" columns => write
# ------------------------------------------------------------------------
def build_master_monthly_report():
    all_files = [f for f in os.listdir(ARCHIVE_FOLDER) if f.lower().endswith(".xlsx")]

    # Identify months
    month_file_info = []
    for fname in all_files:
        match = month_pattern.search(fname)
        if match:
            abbr = match.group(1).capitalize()
            if abbr in month_map:
                sheet_name, sort_idx = month_map[abbr]
                full_path = os.path.join(ARCHIVE_FOLDER, fname)
                month_file_info.append((sort_idx, sheet_name, full_path))

    # Sort by month index (1=Jan, 2=Feb, etc.)
    month_file_info.sort(key=lambda x: x[0])

    # 4.1) Read all data first
    monthly_data_list = []
    for (sort_idx, sheet_name, filepath) in month_file_info:
        df_month = process_monthly_file(filepath)
        monthly_data_list.append((sort_idx, sheet_name, df_month))

    # 4.2) For each month after the first, add two columns per numeric col:
    #      1) "(Qty)" difference
    #      2) "(Percent)" difference
    for i in range(len(monthly_data_list)):
        if i == 0:
            # baseline, no prior data => skip
            continue

        _, sheet_current, df_current = monthly_data_list[i]
        _, sheet_previous, df_previous = monthly_data_list[i - 1]

        # Merge on (RouteName, Day) to line up previous month values
        # Suffix previous columns with _prev
        merged_compare = pd.merge(
            df_current,
            df_previous,
            on=["RouteName", "Day"],
            how="left",
            suffixes=("", "_prev")
        )

        # Fill NaN => 0 for the previous data
        for col in NUMERIC_COLS:
            merged_compare[f"{col}_prev"] = merged_compare[f"{col}_prev"].fillna(0)

        # Build new columns at the end
        new_cols = []
        for col in NUMERIC_COLS:
            # example: # of Services => "Changes from Last Month - # of Services (Qty)"
            qty_colname = f"Changes from Last Month - {col} (Qty)"
            pct_colname = f"Changes from Last Month - {col} (Percent)"
            new_cols.extend([qty_colname, pct_colname])

            # compute difference
            def _compute(row):
                diff, pct = compute_differences(row[col], row[f"{col}_prev"])
                return pd.Series([diff, pct])

            merged_compare[[qty_colname, pct_colname]] = merged_compare.apply(_compute, axis=1)

        # Reorder so new columns appear last
        final_order = list(df_current.columns) + new_cols
        df_current = merged_compare[final_order]

        # Update the stored version
        monthly_data_list[i] = (monthly_data_list[i][0], sheet_current, df_current)

    # 4.3) Build a Summary sheet (Month vs TotalServices) for convenience
    summary_rows = []
    for sort_idx, sheet_name, df_month in monthly_data_list:
        if df_month.empty:
            continue
        try:
            total_services = df_month["# of Services"].sum()
        except Exception:
            total_services = 0
        try:
            total_yardage = df_month["Total Yardage"].sum()
        except Exception:
            total_yardage = 0
        summary_rows.append({
            "Month": sheet_name,
            "TotalServices": total_services,
            "TotalYardage": total_yardage,
        })
    summary_df = pd.DataFrame(summary_rows)

    # 4.4) Write out to Master workbook
    #      Then apply conditional formatting for negative=red, positive=green on difference columns
    with pd.ExcelWriter(MASTER_WORKBOOK_PATH, engine="openpyxl") as writer:
        # Write Summary first (to match Roll-Off behavior)
        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

        for sort_idx, sheet_name, df_month in monthly_data_list:
            if df_month.empty:
                continue

            # Write data to a sheet
            df_month.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # If it's the FIRST month => no "Changes" columns => skip formatting
            if sort_idx == 1:
                # i.e. Jan-2025 is baseline => no changes
                continue

            # We'll find all columns that have either (Qty) or (Percent) in the name
            # and apply conditional formatting
            max_row = len(df_month) + 1  # +1 for header
            from openpyxl.utils import get_column_letter

            # Identify all columns
            for col_index, col_name in enumerate(df_month.columns, start=1):
                if "(Qty)" in col_name or "(Percent)" in col_name:
                    # Build range (e.g. "F2:F100" to skip the header row)
                    start_cell = f"{get_column_letter(col_index)}2"
                    end_cell = f"{get_column_letter(col_index)}{max_row}"
                    cell_range = f"{start_cell}:{end_cell}"

                    # Rule for negative => red
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    rule_neg = CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
                    ws.conditional_formatting.add(cell_range, rule_neg)

                    # Rule for positive => green
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    rule_pos = CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill)
                    ws.conditional_formatting.add(cell_range, rule_pos)

                    # If "(Percent)" in col name => optionally format as a percentage
                    # We'll store it as decimal, so let's apply number format for display
                    if "(Percent)" in col_name:
                        for row_i in range(2, max_row+1):
                            cell = ws.cell(row=row_i, column=col_index)
                            cell.number_format = '0.00%'

    # Also save a copy with latest month suffix, e.g., _Aug2025
    if month_file_info:
        latest_sheet = month_file_info[-1][1]  # e.g., 'Aug-2025'
        suffix = latest_sheet.replace('-', '')  # 'Aug2025'
        suffix = os.environ.get('MASTER_SUFFIX', suffix)
        suffixed_name = f"Master_Monthly_FL_Yardage_Report_Updated_{suffix}.xlsx"
        suffixed_path = os.path.join(MASTER_OUTPUT_FOLDER, suffixed_name)
        try:
            shutil.copyfile(MASTER_WORKBOOK_PATH, suffixed_path)
            print(f"Master (suffixed) also saved at: {suffixed_path}")
        except Exception as e:
            print(f"Warning: could not write suffixed copy: {e}")

    print(f"Master monthly report created at: {MASTER_WORKBOOK_PATH}")
    

# ------------------------------------------------------------------------
# 5) Entry point
# ------------------------------------------------------------------------
if __name__ == "__main__":
    build_master_monthly_report()

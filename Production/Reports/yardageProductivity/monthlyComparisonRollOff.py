import pandas as pd
import os
import re
import numpy as np

# For conditional formatting
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# ────────────────────────────────────────────────────────────────────────────
# 1) Configuration for Roll-Off master comparison
# ────────────────────────────────────────────────────────────────────────────
ARCHIVE_FOLDER = (
    r"C:\Users\arodriguez\Documents\repoProjects\sousan\Production\Reports"
    r"\yardageProductivity\files\Archive\2025\MonthlyReports\RollOff"
)
MASTER_OUTPUT_FOLDER = (
    r"C:\Users\arodriguez\Documents\repoProjects\sousan\Production\Reports"
    r"\yardageProductivity\files\MasterFile\RollOff"
)
MASTER_FILENAME = "Master_Monthly_RollOff_Yardage_Report_Updated.xlsx"

os.makedirs(MASTER_OUTPUT_FOLDER, exist_ok=True)
MASTER_WORKBOOK_PATH = os.path.join(MASTER_OUTPUT_FOLDER, MASTER_FILENAME)

month_pattern = re.compile(
    r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?"
    r"|Aug(?:ust)?|Sep(?:t)?(?:ember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\D?2025",
    re.IGNORECASE
)
month_map = {
    "Jan": ("Jan-2025", 1), "January": ("Jan-2025", 1),
    "Feb": ("Feb-2025", 2), "February": ("Feb-2025", 2),
    "Mar": ("March-2025", 3), "March": ("March-2025", 3),
    "Apr": ("Apr-2025", 4), "April": ("Apr-2025", 4),
    "May": ("May-2025", 5),
    "Jun": ("Jun-2025", 6), "June": ("Jun-2025", 6),
    "Jul": ("Jul-2025", 7), "July": ("Jul-2025", 7),
    "Aug": ("Aug-2025", 8), "August": ("Aug-2025", 8),
    "Sep": ("Sep-2025", 9), "Sept": ("Sep-2025", 9), "September": ("Sep-2025", 9),
    "Oct": ("Oct-2025", 10), "October": ("Oct-2025", 10),
    "Nov": ("Nov-2025", 11), "November": ("Nov-2025", 11),
    "Dec": ("Dec-2025", 12), "December": ("Dec-2025", 12),
}

# roll-off container sizes
ROLL_OFF_SIZES = [10, 15, 20, 30, 40]

# metrics we'll compare month-over-month
NUMERIC_COLS = ["# of Services", "Total Yardage"] + [f"Size-{sz}" for sz in ROLL_OFF_SIZES]
BASE_COLS    = ["RouteName", "Day"] + NUMERIC_COLS

# for weekday sorting
day_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
day_index = {d:i for i,d in enumerate(day_order)}


# ────────────────────────────────────────────────────────────────────────────
# 2) Read & summarize each monthly Roll-Off file
# ────────────────────────────────────────────────────────────────────────────
def process_monthly_file(path: str) -> pd.DataFrame:
    try:
        df_all = pd.read_excel(path, sheet_name="All Data", engine="openpyxl")
    except Exception:
        return pd.DataFrame(columns=BASE_COLS)
    if df_all.empty:
        return pd.DataFrame(columns=BASE_COLS)

    svc    = df_all.groupby(["RouteName","Day"]).size().reset_index(name="# of Services")
    yard   = df_all.groupby(["RouteName","Day"])["Yardage"].sum().reset_index(name="Total Yardage")
    merged = pd.merge(svc, yard, on=["RouteName","Day"], how="outer")

    try:
        df_sum = pd.read_excel(path, sheet_name="Summary by Route-Day", engine="openpyxl")
        pivot  = (
            df_sum
            .pivot_table(index=["RouteName","Day"],
                         columns="Size",
                         values="Count",
                         aggfunc="sum",
                         fill_value=0)
            .reset_index()
        )
    except Exception:
        pivot = pd.DataFrame(columns=["RouteName","Day"])

    # rename pivot cols to "Size-<n>"
    rename_map = {col: f"Size-{int(col)}" for col in pivot.columns if isinstance(col,(int,float))}
    pivot.rename(columns=rename_map, inplace=True)
    # ensure every roll-off size is present
    for sz in ROLL_OFF_SIZES:
        col = f"Size-{sz}"
        if col not in pivot.columns:
            pivot[col] = 0

    df = pd.merge(merged, pivot, on=["RouteName","Day"], how="outer")
    for c in NUMERIC_COLS:
        df[c] = df.get(c,0).fillna(0)

    df["__sort__"] = df.apply(
        lambda r: (
            0 if "Route" in r["RouteName"] else 1,
            int(re.search(r"(\d+)$", r["RouteName"]).group(1)) if re.search(r"(\d+)$", r["RouteName"]) else 999,
            day_index.get(r["Day"],99)
        ), axis=1
    )
    df = df.sort_values("__sort__").drop(columns="__sort__")

    return df[BASE_COLS]


# ────────────────────────────────────────────────────────────────────────────
# 3) Build the Master Roll-Off workbook
# ────────────────────────────────────────────────────────────────────────────
def build_master_monthly_report():
    files = [f for f in os.listdir(ARCHIVE_FOLDER) if f.lower().endswith(".xlsx")]
    month_info = []
    for fname in files:
        m = month_pattern.search(fname)
        if not m: continue
        key = m.group(1).capitalize()
        if key in month_map:
            sheet, idx = month_map[key]
            month_info.append((idx, sheet, os.path.join(ARCHIVE_FOLDER, fname)))
    month_info.sort(key=lambda x: x[0])

    monthly = [(idx, sheet, process_monthly_file(path))
               for idx, sheet, path in month_info]

    # Summary sheet: true total rolls each month
    summary_rows = []
    for _, s, df in monthly:
        if df is None or df.empty:
            summary_rows.append({"Month": s, "TotalServices": 0, "TotalYardage": 0})
        else:
            summary_rows.append({
                "Month": s,
                "TotalServices": df["# of Services"].sum(),
                "TotalYardage": df["Total Yardage"].sum(),
            })
    summary_df = pd.DataFrame(summary_rows)

    # month-over-month differences
    for i in range(1, len(monthly)):
        _, sheet_cur, df_cur = monthly[i]
        _, _, df_prev       = monthly[i-1]

        merged = pd.merge(df_cur, df_prev,
                          on=["RouteName","Day"],
                          how="left",
                          suffixes=("", "_prev"))
        for col in NUMERIC_COLS:
            merged[f"{col}_prev"] = merged[f"{col}_prev"].fillna(0)

        for col in NUMERIC_COLS:
            diff = merged[col] - merged[f"{col}_prev"]
            prev = merged[f"{col}_prev"]
            pct  = diff.div(prev).where(
                prev != 0,
                pd.Series(
                    np.where(diff>0, 1.0,
                             np.where(diff<0, -1.0, 0.0)),
                    index=diff.index
                )
            )
            merged[f"Changes - {col} (Qty)"] = diff
            merged[f"Changes - {col} (%)"]    = pct

        order = (
            list(df_cur.columns)
            + [f"Changes - {c} (Qty)" for c in NUMERIC_COLS]
            + [f"Changes - {c} (%)"   for c in NUMERIC_COLS]
        )
        monthly[i] = (monthly[i][0], sheet_cur, merged[order])

    # write & apply conditional formatting
    with pd.ExcelWriter(MASTER_WORKBOOK_PATH, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # define our fills
        red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for idx, sheet, df in monthly:
            df.to_excel(writer, sheet_name=sheet, index=False)
            if idx == monthly[0][0]:
                continue

            ws = writer.sheets[sheet]
            max_row = len(df) + 1
            from openpyxl.utils import get_column_letter

            for ci, cname in enumerate(df.columns, start=1):
                if cname.startswith("Changes - "):
                    colL = get_column_letter(ci)
                    rng  = f"{colL}2:{colL}{max_row}"

                    # negative → red
                    ws.conditional_formatting.add(
                        rng,
                        CellIsRule(operator="lessThan",
                                   formula=["0"],
                                   fill=red_fill)
                    )
                    # positive → green
                    ws.conditional_formatting.add(
                        rng,
                        CellIsRule(operator="greaterThan",
                                   formula=["0"],
                                   fill=green_fill)
                    )
                    if cname.endswith("(%)"):
                        for r in range(2, max_row+1):
                            ws.cell(row=r, column=ci).number_format = "0.00%"

    print(f"✅ Master Roll-Off report created at:\n   {MASTER_WORKBOOK_PATH}")
    # Also save a copy with latest month suffix, e.g., _Aug2025
    if month_info:
        latest_sheet = month_info[-1][1]  # e.g., 'Aug-2025'
        suffix = latest_sheet.replace('-', '')  # 'Aug2025'
        suffix = os.environ.get('MASTER_SUFFIX', suffix)
        suffixed_name = f"Master_Monthly_RollOff_Yardage_Report_Updated_{suffix}.xlsx"
        suffixed_path = os.path.join(MASTER_OUTPUT_FOLDER, suffixed_name)
        try:
            import shutil
            shutil.copyfile(MASTER_WORKBOOK_PATH, suffixed_path)
            print(f"Master (suffixed) also saved at:\n   {suffixed_path}")
        except Exception as e:
            print(f"Warning: could not write suffixed copy: {e}")

if __name__ == "__main__":
    build_master_monthly_report()
